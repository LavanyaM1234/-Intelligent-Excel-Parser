"""Generate test Excel files for the parser"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import os
from pathlib import Path
from datetime import datetime, timedelta
import random


def create_clean_data():
    """
    Create clean_data.xlsx: Clean headers, single asset, numeric values
    Baseline — should parse perfectly
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Factory Data"
    
    # Header row
    headers = ["Date", "Coal Consumption", "Steam Generation", "Power Generation", "Efficiency"]
    ws.append(headers)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data rows with clean numeric values
    data = [
        ["2024-01-01", 500.0, 250.5, 125.3, 0.85],
        ["2024-01-02", 520.0, 260.0, 130.0, 0.87],
        ["2024-01-03", 480.5, 240.0, 120.5, 0.83],
        ["2024-01-04", 510.0, 255.0, 128.0, 0.86],
        ["2024-01-05", 495.0, 248.0, 124.0, 0.84],
    ]
    
    for row in data:
        ws.append(row)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 20
    
    wb.save("data/clean_data.xlsx")
    print("✓ Created data/clean_data.xlsx")


def create_messy_data():
    """
    Create messy_data.xlsx: Abbreviated headers, mixed formats, title rows, empty rows
    Fuzzy matching + header detection
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Operations"
    
    # Title row (should be skipped)
    ws.append(["Factory Monthly Report - January 2024"])
    ws.append([])  # Empty row
    
    # Actual header row (should be detected as row 3)
    headers = ["Date", "Coal Used (MT)", "Steam (Boiler 1)", "POWER GEN", "Overall Efficiency %"]
    ws.append(headers)
    
    # Style header
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data rows with messy formats
    data = [
        ["2024-01-01", "1,200.50", "580.25", "290", "85%"],
        ["2024-01-02", "1220.00", "600", "300.5", "87%"],
        ["2024-01-03", None, "570.5", "285", "N/A"],  # Contains None
        ["2024-01-04", "1,210", "590.75", "295 MW", "86%"],  # Mixed format
        ["2024-01-05", "1,180.75", "565.00", "280", "84%"],
        [],  # Empty row
        ["2024-01-06", "YES", "TRUE", "1", "0.88"],  # Boolean-like values
    ]
    
    for row in data:
        ws.append(row)
    
    wb.save("data/messy_data.xlsx")
    print("✓ Created data/messy_data.xlsx")


def create_multi_asset_data():
    """
    Create multi_asset.xlsx: Column headers include asset names, multiple assets per parameter
    Asset detection + parameter deduplication
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Multi-Asset"
    
    # Header row with asset names embedded
    headers = [
        "Date",
        "Coal Consumption AFBC-1",
        "Coal Consumption AFBC-2",
        "Steam (Boiler 1)",
        "Power TG-1",
        "Power TG-2",
        "Efficiency",
    ]
    ws.append(headers)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data rows
    data = [
        ["2024-01-01", 500.0, 450.0, 250.5, 125.0, 120.0, 85],
        ["2024-01-02", 520.0, 480.0, 260.0, 130.0, 125.0, 87],
        ["2024-01-03", 480.0, 470.0, 240.0, 120.0, 118.0, 83],
        ["2024-01-04", 510.0, 490.0, 255.0, 128.0, 122.0, 86],
        ["2024-01-05", 495.0, 465.0, 248.0, 124.0, 119.0, 84],
    ]
    
    for row in data:
        ws.append(row)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 25
    
    wb.save("data/multi_asset_data.xlsx")
    print("✓ Created data/multi_asset_data.xlsx")


def create_headers_only_data():
    """
    Create headers_only.xlsx: Only header row, no data
    Tests: Empty data handling, warnings about no rows
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Headers Only"
    
    # Only header row, no data
    headers = ["Date", "Coal Consumption", "Steam Generation", "Power Generation", "Efficiency"]
    ws.append(headers)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    wb.save("data/headers_only_data.xlsx")
    print("✓ Created data/headers_only_data.xlsx")


def create_sparse_data():
    """
    Create sparse_data.xlsx: Columns with mostly null/N/A values
    Tests: Null value handling, sparse data validation
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sparse Data"
    
    # Header row
    headers = ["Date", "Coal Consumption", "Steam Generation", "Power Generation", "Comments", "Efficiency"]
    ws.append(headers)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data rows with sparse values
    data = [
        ["2024-01-01", 500.0, None, None, "Calibration day", None],
        ["2024-01-02", None, 260.0, None, "Maintenance", "N/A"],
        ["2024-01-03", None, None, 120.5, "Test run", None],
        ["2024-01-04", "N/A", "N/A", "N/A", "System offline", "N/A"],
        ["2024-01-05", 495.0, None, None, None, 0.84],
    ]
    
    for row in data:
        ws.append(row)
    
    wb.save("data/sparse_data.xlsx")
    print("✓ Created data/sparse_data.xlsx")


def create_duplicate_headers():
    """
    Create duplicate_headers.xlsx: Same parameter/asset in multiple columns
    Tests: Column duplication detection and warnings
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Duplicates"
    
    # Header row with duplicate columns
    headers = [
        "Date",
        "Coal Consumption",  # First occurrence
        "Efficiency",
        "Coal Consumption",  # Duplicate (should warn)
        "Steam Generation",
        "Power Generation (MW)",
        "Power Generation",  # Another potential duplicate (MW vs none)
    ]
    ws.append(headers)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data rows
    data = [
        ["2024-01-01", 500.0, 85, 505.0, 250.5, 125.0, 125.5],
        ["2024-01-02", 520.0, 87, 515.0, 260.0, 130.0, 130.0],
        ["2024-01-03", 480.0, 83, 485.0, 240.0, 120.5, 120.0],
    ]
    
    for row in data:
        ws.append(row)
    
    wb.save("data/duplicate_headers.xlsx")
    print("✓ Created data/duplicate_headers.xlsx")


def create_special_chars_headers():
    """
    Create special_chars_headers.xlsx: Headers with special characters, mixed case
    Tests: Header normalization, fuzzy matching with special chars
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Special Chars"
    
    # Header row with special characters and mixed case
    headers = [
        "Date (YYYY-MM-DD)",
        "Coal Consumption [MT]",
        "STEAM_GENERATION (T/hr)",
        "Power Output % (MWh)",
        "#Efficiency_Ratio",
        "Factory_ID/Asset",
    ]
    ws.append(headers)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data rows
    data = [
        ["2024-01-01", 500.0, 250.5, 125.3, 0.85, "AFBC-1"],
        ["2024-01-02", 520.0, 260.0, 130.0, 0.87, "AFBC-2"],
        ["2024-01-03", 480.5, 240.0, 120.5, 0.83, "TG-1"],
    ]
    
    for row in data:
        ws.append(row)
    
    wb.save("data/special_chars_headers.xlsx")
    print("✓ Created data/special_chars_headers.xlsx")


def create_validation_errors():
    """
    Create validation_errors.xlsx: Values that fail validation rules
    Tests: Validation warnings, suspicious values detection
    - Negative coal consumption
    - Efficiency > 100%
    - Power values with mixed units
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Bad Data"
    
    # Header row
    headers = ["Date", "Coal Consumption", "Efficiency", "Power Generation", "Steam Generation"]
    ws.append(headers)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data rows with validation issues
    data = [
        ["2024-01-01", -500.0, 0.85, 125.0, 250.5],  # Negative coal
        ["2024-01-02", 520.0, 150, 130.0, 260.0],  # Efficiency > 100%
        ["2024-01-03", 480.5, 0.83, -120.5, 240.0],  # Negative power
        ["2024-01-04", 510.0, 1.05, 128.0, 255.0],  # Efficiency > 100% (decimal)
        ["2024-01-05", 0, 0.84, 0, 248.0],  # Zero values (valid but edge case)
    ]
    
    for row in data:
        ws.append(row)
    
    wb.save("data/validation_errors.xlsx")
    print("✓ Created data/validation_errors.xlsx")


def create_mixed_formats_data():
    """
    Create mixed_formats.xlsx: Extreme variety of number formats
    Tests: Value parsing robustness across formats
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Mixed Formats"
    
    # Header row
    headers = ["Date", "Coal Consumption", "Efficiency", "Steam Generation", "Notes"]
    ws.append(headers)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data rows with mixed format values
    data = [
        ["2024-01-01", "500", "85%", "250.5", "Plain strings"],
        ["2024-01-02", "1,200.50", "87 %", "260", "Comma decimals and space"],
        ["2024-01-03", "480.5 MT", "0.83", "240 T/hr", "With units"],
        ["2024-01-04", "510", "86%", "255", "Clean mixed"],
        ["2024-01-05", "495.00", "84 percent", "248", "Word format"],
        ["2024-01-06", "500 MT/day", "0.86", "250", "Complex units"],
    ]
    
    for row in data:
        ws.append(row)
    
    wb.save("data/mixed_formats_data.xlsx")
    print("✓ Created data/mixed_formats_data.xlsx")


def create_large_data_bulk():
    """
    Create large_data_5mb.xlsx: Large file (~5MB) with many rows
    Tests: Large file handling, chunking capability
    Generates ~10,000 rows to reach ~5MB file size
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Bulk Data"
    
    # Header row
    headers = ["Date", "Coal Consumption", "Steam Generation", "Power Generation", "Efficiency"]
    ws.append(headers)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Generate many rows of data
    base_date = datetime(2023, 1, 1)
    num_rows = 15000  # Generate 15k rows for a ~5MB file
    
    print(f"  Generating {num_rows:,} rows for large file...")
    for i in range(num_rows):
        date = base_date + timedelta(hours=i)
        coal = random.uniform(400, 600)
        steam = random.uniform(200, 300)
        power = random.uniform(100, 150)
        efficiency = random.uniform(0.80, 0.95)
        
        ws.append([date.strftime("%Y-%m-%d %H:%M"), coal, steam, power, efficiency])
        
        # Print progress every 3000 rows
        if (i + 1) % 3000 == 0:
            print(f"    {i + 1:,} rows written...")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 20
    
    wb.save("data/large_data_5mb.xlsx")
    file_size_mb = os.path.getsize("data/large_data_5mb.xlsx") / (1024 * 1024)
    print(f"✓ Created data/large_data_5mb.xlsx ({file_size_mb:.2f} MB, {num_rows:,} rows)")


if __name__ == "__main__":
    # Create data directory if it doesn't exist
    Path("data").mkdir(exist_ok=True)
    
    print("Generating test Excel files for edge cases...\n")
    print("=" * 60)
    
    print("\n1. BASELINE - Clean data:")
    create_clean_data()
    
    print("\n2. MESSY FORMAT - Headers, title rows, empty rows, mixed formats:")
    create_messy_data()
    
    print("\n3. MULTI-ASSET - Multiple assets per parameter:")
    create_multi_asset_data()
    
    print("\n4. HEADERS ONLY - No data rows:")
    create_headers_only_data()
    
    print("\n5. SPARSE DATA - Mostly null/N/A values:")
    create_sparse_data()
    
    print("\n6. DUPLICATE HEADERS - Same parameter in multiple columns:")
    create_duplicate_headers()
    
    print("\n7. SPECIAL CHARACTERS - Headers with special chars & mixed case:")
    create_special_chars_headers()
    
    print("\n8. VALIDATION ERRORS - Negative values, efficiency > 100%:")
    create_validation_errors()
    
    print("\n9. MIXED FORMATS - Various number formats and units:")
    create_mixed_formats_data()
    
    print("\n10. LARGE FILE (~5MB) - Many rows for chunking test:")
    create_large_data_bulk()
    
    print("\n" + "=" * 60)
    print("\n✓ All test files created successfully!")
    print("\nTest file summary:")
    print("  - clean_data.xlsx: Baseline test")
    print("  - messy_data.xlsx: Header detection & fuzzy matching")
    print("  - multi_asset_data.xlsx: Asset detection")
    print("  - headers_only_data.xlsx: Empty file handling")
    print("  - sparse_data.xlsx: Null/N/A handling")
    print("  - duplicate_headers.xlsx: Duplicate detection")
    print("  - special_chars_headers.xlsx: Special char normalization")
    print("  - validation_errors.xlsx: Validation rule checking")
    print("  - mixed_formats_data.xlsx: Format robustness")
    print("  - large_data_5mb.xlsx: Large file & chunking (~5MB)")
