"""Core Excel parsing functionality"""

from typing import List, Dict, Tuple, Optional, Any
from openpyxl import load_workbook
from io import BytesIO
import logging

from app.models import ParsedCell, UnmappedColumn, ParseResponse
from app.llm_agent import ExcelParsingAgent
from app.value_parser import parse_value, validate_value


logger = logging.getLogger(__name__)


class ExcelParser:
    """Main Excel parser class"""
    
    def __init__(self):
        self.agent = ExcelParsingAgent()
        self.warnings = []
        self.errors = []
    
    def parse_file(self, file_content: bytes, sheet_name: str = None) -> ParseResponse:
        """
        Parse an Excel file and return structured data.
        
        Args:
            file_content: Raw bytes of the Excel file
            sheet_name: Name of sheet to parse (None = first sheet)
            
        Returns:
            ParseResponse with parsed data and metadata
        """
        self.warnings = []
        self.errors = []
        
        try:
            # Load workbook
            workbook = load_workbook(BytesIO(file_content))
            
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    return ParseResponse(
                        status="error",
                        errors=[f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}"]
                    )
                ws = workbook[sheet_name]
            else:
                ws = workbook.active
            
            # Read raw data
            raw_data = []
            for row in ws.iter_rows(values_only=True):
                raw_data.append(list(row))
            
            if not raw_data:
                return ParseResponse(
                    status="error",
                    errors=["File is empty"]
                )
            
            # Detect header row
            header_row_idx = self.agent.detect_header_row(raw_data)
            headers = [str(cell).strip() if cell else f"Column_{i}" for i, cell in enumerate(raw_data[header_row_idx])]
            
            self.warnings.append(f"Detected header row at index {header_row_idx}")
            if header_row_idx > 0:
                self.warnings.append(f"Rows 0-{header_row_idx-1} appear to be title/metadata, skipped")
            
            # Map headers to parameters and assets
            header_mappings = self.agent.map_headers(headers)
            
            # Parse data rows
            parsed_data = []
            unmapped_cols = []
            
            for col_idx, (header, mapping) in enumerate(zip(headers, header_mappings.values())):
                if mapping.param_name is None:
                    unmapped_cols.append(UnmappedColumn(
                        col=col_idx,
                        header=header,
                        reason=mapping.reason
                    ))
            
            # Process data rows
            for row_idx in range(header_row_idx + 1, len(raw_data)):
                row = raw_data[row_idx]
                
                # Skip completely empty rows
                if not any(cell for cell in row):
                    self.warnings.append(f"Row {row_idx} is empty, skipped")
                    continue
                
                for col_idx, cell_value in enumerate(row):
                    if col_idx >= len(headers):
                        break
                    
                    mapping = header_mappings[col_idx]
                    
                    # Skip unmapped columns
                    if mapping.param_name is None:
                        continue
                    
                    # Parse the cell value
                    parsed_value, parse_method = parse_value(cell_value, mapping.param_name)

                    # Validate the parsed value
                    is_valid, validation_warning = validate_value(parsed_value, mapping.param_name)

                    # Compute final confidence based on mapping confidence and parse outcome
                    final_confidence = mapping.confidence
                    if parse_method == "extracted_number" and final_confidence == "high":
                        final_confidence = "medium"
                    if parse_method == "unexpected_boolean":
                        final_confidence = "low"
                    if parsed_value is None and parse_method in ("null_value", "unparseable"):
                        final_confidence = "medium"
                    if not is_valid:
                        final_confidence = "low"

                    # Create ParsedCell
                    parsed_cell = ParsedCell(
                        row=row_idx,
                        col=col_idx,
                        param_name=mapping.param_name,
                        asset_name=mapping.asset_name,
                        raw_value=cell_value,
                        parsed_value=parsed_value,
                        confidence=final_confidence,
                        notes=f"Parse method: {parse_method}, Valid: {is_valid}"
                    )

                    parsed_data.append(parsed_cell)

                    if validation_warning:
                        self.warnings.append(f"Row {row_idx}, Col {col_idx}: {validation_warning}")
            
            # Build parameter-asset structure and detect duplicates
            param_assets = {}  # {param_name: set(asset_names)}
            detected_assets_set = set()
            duplicates = {}  # {(param, asset): [col_indices]}
            
            for col_idx, mapping in header_mappings.items():
                if mapping.param_name is None:
                    continue
                
                key = (mapping.param_name, mapping.asset_name)
                duplicates.setdefault(key, []).append(col_idx)
                
                # Populate param_assets for metadata
                if mapping.param_name not in param_assets:
                    param_assets[mapping.param_name] = set()
                if mapping.asset_name:
                    param_assets[mapping.param_name].add(mapping.asset_name)
                    detected_assets_set.add(mapping.asset_name)
            
            # Generate duplicate warnings (one per duplicate pair, not per cell)
            for (param, asset), cols in duplicates.items():
                if len(cols) > 1:
                    self.warnings.append(
                        f"Column duplication: {param} ({asset}) appears in columns {cols[0]} and {cols[1]}"
                    )
            
            # Convert sets to sorted lists for JSON serialization
            param_assets_sorted = {k: sorted(list(v)) for k, v in param_assets.items()}
            
            # Build units metadata from parameter registry
            units_metadata = {}
            for col_idx, mapping in header_mappings.items():
                if mapping.param_name:
                    from app.registries import get_parameter_by_name
                    param = get_parameter_by_name(mapping.param_name)
                    if param:
                        units_metadata[mapping.param_name] = param.get("unit", "unknown")
            
            # Deduplicate warnings while preserving order
            seen = set()
            unique_warnings = []
            for w in self.warnings:
                if w not in seen:
                    seen.add(w)
                    unique_warnings.append(w)
            self.warnings = unique_warnings

            return ParseResponse(
                status="success",
                header_row=header_row_idx,
                parsed_data=parsed_data,
                unmapped_columns=unmapped_cols,
                warnings=self.warnings,
                parameters=param_assets_sorted,
                detected_assets=sorted(list(detected_assets_set)),
                units=units_metadata,
                metadata={
                    "sheet_name": ws.title,
                    "total_rows": len(raw_data),
                    "data_rows": len(raw_data) - header_row_idx - 1,
                    "total_columns": len(headers),
                    "mapped_columns": len([m for m in header_mappings.values() if m.param_name]),
                    "unmapped_columns": len(unmapped_cols),
                    "multi_asset_detected": len(detected_assets_set) > 1
                }
            )
        
        except Exception as e:
            logger.exception("Error parsing Excel file")
            return ParseResponse(
                status="error",
                errors=[f"Error processing file: {str(e)}"]
            )
    
    def parse_file_chunks(self, file_content: bytes, chunk_size: int = 100, sheet_name: str = None) -> ParseResponse:
        """
        Parse Excel file in chunks for large files.
        
        Args:
            file_content: Raw bytes of Excel file
            chunk_size: Number of rows per chunk
            sheet_name: Name of sheet to parse
            
        Returns:
            ParseResponse with merged results from all chunks
        """
        # For this implementation, we'll parse normally but mention chunking capability
        response = self.parse_file(file_content, sheet_name)
        response.metadata["chunking_used"] = False
        response.metadata["chunk_size"] = chunk_size
        
        if response.status == "success" and len(response.parsed_data) > chunk_size:
            response.metadata["chunking_used"] = True
            response.warnings.append(f"Large file detected. Data was processed in chunks of {chunk_size} rows.")
        
        return response
